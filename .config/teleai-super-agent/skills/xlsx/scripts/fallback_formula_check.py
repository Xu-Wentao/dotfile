#!/usr/bin/env python3
"""
Degraded-mode fallback checks for Excel formula validation.

Normal mode:
  - `python recalc.py <excel_file>` uses LibreOffice (soffice) to recalc formulas
    and then scans cached formula results for Excel error strings.

Degraded mode (when LibreOffice/soffice is unavailable and cannot be installed):
  - This script performs best-effort checks using pure Python + openpyxl:
    1) Scan cached values (data_only=True) for error strings like `#REF!`.
    2) Parse formula text (data_only=False) to detect obviously invalid
       cross-sheet references and out-of-bound cell addresses.

Important limitations:
  - Without LibreOffice recalculation, `data_only=True` may not contain
    up-to-date formula results (Excel cached values may be missing/empty).
  - Static formula parsing cannot guarantee that the formula will evaluate
    correctly in Excel; it only checks basic reference structure.

Output:
  - A JSON report (to `--out` or stdout) with clear `[DEGRADED MODE]` messaging.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Error: openpyxl is required for fallback_formula_check.\n"
        "Run: pip install openpyxl"
    ) from exc


EXCEL_ERROR_TOKENS: tuple[str, ...] = (
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)

MAX_EXCEL_COLS = 16384  # XFD
MAX_EXCEL_ROWS = 1048576


SHEET_NAME_RE = r"(?:'[^']+'|[A-Za-z0-9_\.]+)"
CELL_ADDR_RE = r"(?:\$?[A-Z]{1,3}\$?\d+)"
CELL_REF_RE = re.compile(
    rf"(?:(?P<sheet>{SHEET_NAME_RE})!)?(?P<addr>{CELL_ADDR_RE})",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True)
class ParsedCell:
    sheet: str | None
    col: int
    row: int
    raw: str


def _excel_col_to_number(col_letters: str) -> int | None:
    """
    Convert Excel column letters to 1-indexed column number.

    Args:
        col_letters: e.g. "A", "BC" (may include '$' which will be ignored).

    Returns:
        Column number, or None if invalid.
    """

    letters = col_letters.replace("$", "").upper()
    if not letters or not re.fullmatch(r"[A-Z]{1,3}", letters):
        return None

    col_num = 0
    for ch in letters:
        col_num = col_num * 26 + (ord(ch) - ord("A") + 1)
    return col_num


def _parse_cell_address(addr: str) -> tuple[int, int] | None:
    """
    Parse an A1-style cell address into (col_num, row_num).
    Accepts absolute markers `$` like `$A$1`.
    """

    m = re.fullmatch(r"(\$?[A-Z]{1,3})(\$?)(\d+)", addr, flags=re.IGNORECASE)
    if not m:
        return None
    col_letters = m.group(1)
    row_str = m.group(3)
    col_num = _excel_col_to_number(col_letters)
    if col_num is None:
        return None
    row_num = int(row_str)
    return col_num, row_num


def _normalize_sheet_name(sheet: str) -> str:
    """
    Normalize a sheet name possibly wrapped in single quotes.
    """

    s = sheet.strip()
    if len(s) >= 2 and s[0] == "'" and s[-1] == "'":
        return s[1:-1]
    return s


def _scan_cached_value_errors(excel_path: Path, *, max_locations: int) -> dict[str, Any]:
    """
    Scan cached values for Excel error tokens using data_only=True.
    """

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        error_details: dict[str, list[str]] = {t: [] for t in EXCEL_ERROR_TOKENS}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    for token in EXCEL_ERROR_TOKENS:
                        if token in val:
                            loc = f"{sheet_name}!{cell.coordinate}"
                            error_details[token].append(loc)
                            total_errors += 1
                            break

        summary: dict[str, Any] = {}
        for token, locs in error_details.items():
            if not locs:
                continue
            summary[token] = {
                "count": len(locs),
                "locations": locs[:max_locations],
            }

        return {
            "cached_value_errors_found": total_errors > 0,
            "total_cached_value_errors": total_errors,
            "cached_error_summary": summary,
        }
    finally:
        wb.close()


def _static_check_formula_references(
    excel_path: Path,
    *,
    max_formula_cells: int,
) -> dict[str, Any]:
    """
    Best-effort parse formula text for basic reference sanity.

    Checks:
      - referenced sheet names exist (when explicitly qualified)
      - referenced cell addresses are within Excel absolute bounds
    """

    wb = load_workbook(excel_path, data_only=False, read_only=True)
    try:
        sheet_set = set(wb.sheetnames)
        out_of_bounds: list[str] = []
        missing_sheets: set[str] = set()

        formula_cell_scanned = 0
        unknown_patterns_notes: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    if not val.startswith("="):
                        continue
                    formula_cell_scanned += 1
                    if formula_cell_scanned > max_formula_cells:
                        unknown_patterns_notes.append(
                            f"Scanned formula cells capped at {max_formula_cells}; "
                            "static checks may be incomplete."
                        )
                        return {
                            "static_checks_complete": False,
                            "static_reference_issues": {
                                "missing_sheets": sorted(missing_sheets),
                                "out_of_bound_references": out_of_bounds[:200],
                            },
                            "static_notes": unknown_patterns_notes,
                        }

                    formula_text = val
                    # Cheap heuristic: structured references often contain '@' or '{' patterns.
                    if "{" in formula_text or "@" in formula_text:
                        unknown_patterns_notes.append(
                            "Detected structured reference patterns; static checks are best-effort."
                        )

                    for m in CELL_REF_RE.finditer(formula_text):
                        sheet_raw = m.group("sheet")
                        addr_raw = m.group("addr")
                        parsed = _parse_cell_address(addr_raw)
                        if parsed is None:
                            continue
                        col_num, row_num = parsed

                        if sheet_raw is not None:
                            sheet_name_norm = _normalize_sheet_name(sheet_raw)
                            if sheet_name_norm not in sheet_set:
                                missing_sheets.add(sheet_name_norm)
                        # Unqualified refs default to current sheet; no missing-sheet check.

                        if not (1 <= col_num <= MAX_EXCEL_COLS and 1 <= row_num <= MAX_EXCEL_ROWS):
                            out_of_bounds.append(f"{sheet_name}!{cell.coordinate} -> {addr_raw}")

        return {
            "static_checks_complete": True,
            "static_reference_issues": {
                "missing_sheets": sorted(missing_sheets),
                "out_of_bound_references": out_of_bounds[:200],
            },
            "static_notes": unknown_patterns_notes,
        }
    finally:
        wb.close()


def _build_report(
    *,
    excel_path: Path,
    max_locations: int,
    max_formula_cells: int,
) -> dict[str, Any]:
    """
    Build the final JSON report.
    """

    print("[DEGRADED MODE] Running pure-Python Excel fallback checks.", file=sys.stderr)
    print(
        "For authoritative validation (including correct formula evaluation), install LibreOffice/soffice and run `python recalc.py ...`.",
        file=sys.stderr,
    )

    cached = _scan_cached_value_errors(excel_path, max_locations=max_locations)
    static = _static_check_formula_references(excel_path, max_formula_cells=max_formula_cells)

    guarantee = {
        "zero_formula_errors": "cannot_be_guaranteed_without_libreoffice_recalc",
    }

    report: dict[str, Any] = {
        "degraded_mode": True,
        "input_file": str(excel_path),
        "guarantee": guarantee,
        "cached_value_errors_found": cached["cached_value_errors_found"],
        "total_cached_value_errors": cached["total_cached_value_errors"],
        "cached_error_summary": cached["cached_error_summary"],
        "static_checks": static,
        "recommend_install_and_rerun": [
            "Install LibreOffice/soffice (normal mode).",
            "Then rerun: `python recalc.py <excel_file>`.",
        ],
    }
    return report


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fallback Excel formula checks (degraded mode).")
    parser.add_argument("input_excel", type=Path, help="Path to an .xlsx file.")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Write JSON report to this path. If omitted, print to stdout.",
    )
    parser.add_argument("--max-locations", type=int, default=20, help="Max locations per error type.")
    parser.add_argument(
        "--max-formula-cells",
        type=int,
        default=20000,
        help="Cap static formula scanning for performance.",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = _parse_args(argv)
    excel_path = args.input_excel

    if not excel_path.exists():
        print(f"Error: file not found: {excel_path}", file=sys.stderr)
        return 2

    report = _build_report(
        excel_path=excel_path,
        max_locations=args.max_locations,
        max_formula_cells=args.max_formula_cells,
    )

    if args.out is None:
        sys.stdout.write(json.dumps(report, indent=2))
        sys.stdout.write("\n")
        return 0

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"Written report to: {args.out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

